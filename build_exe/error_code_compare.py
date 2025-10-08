import os
import sys
import pandas as pd
from pathlib import Path
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import tkinter.font as tkfont
import tkinter as tk
from tkinter import ttk
from guide_popup.guide import show_guide

# 讀取 UI 文字設定
SETUP_FILE = 'setup.txt'
def load_ui_text():
    default_text = {
        'ErrorCodeXMLLabel': '錯誤碼 XML 檔案：',
        'SourceExcelLabel': '來源 Excel 檔案：',
        'SelectSheetLabel': '選擇工作表：',
        'BrowseButton': '瀏覽',
        'CompareButton': '比對',
        'SuccessTitle': '成功',
        'SuccessMsg': '比對完成，結果已儲存於：',
        'CancelTitle': '取消',
        'CancelMsg': '已取消操作。',
        'FileExistsTitle': '檔案已存在',
        'FileExistsMsg': '檔案 {output_path} 已存在，是否要覆蓋？',
        'ErrorTitle': '錯誤',
        'SheetLoadError': '載入工作表時發生錯誤：{error}',
        'CompareError': '比對時發生錯誤：{error}',
        'NotFound': '查無說明',
        'NotFoundCN': '查無中文說明',
        'WindowWidth': '540',
        'WindowHeight': '340',
        'FontSize': '11',
        'BrowseXMLTooltip': '選擇錯誤碼XML檔案',
        'BrowseExcelTooltip': '選擇來源Excel檔案',
        'BrowseSheetTooltip': '選擇工作表',
    }
    if not os.path.exists(SETUP_FILE):
        with open(SETUP_FILE, 'w', encoding='utf-8') as f:
            for k, v in default_text.items():
                f.write(f'{k}={v}\n')
        return default_text
    ui_text = {}
    with open(SETUP_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            if '=' in line:
                k, v = line.strip().split('=', 1)
                ui_text[k] = v
    for k, v in default_text.items():
        if k not in ui_text:
            ui_text[k] = v
    return ui_text

ui_text = load_ui_text()

# Tooltip class for Tkinter widgets
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)
    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert") if hasattr(self.widget, 'bbox') else (0,0,0,0)
        x = x + self.widget.winfo_rootx() + 30
        y = y + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left', background="#ffffe0", relief='solid', borderwidth=1, font=("Calibri", 10))
        label.pack(ipadx=4)
    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class ErrorCodeComparer:
    def __init__(self, root):
        self.root = root
        self.root.title("Error Code Comparer")
        self.excel1_path = None
        self.excel2_path = None
        self.selected_sheet = None
        # 讀取字體大小
        self.font_size = int(ui_text.get('FontSize', 10))
        self.create_widgets()
        # 讀取寬高，元件繪製完畢後再設定 geometry
        width = int(ui_text.get('WindowWidth', 700))
        height = int(ui_text.get('WindowHeight', 400))
        self.root.update_idletasks()
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(True, True)
        self.center_window()
        self.root.configure(bg="#e6f2ff")  # 設定主視窗背景為淺藍色
        
        # 在初始化完成後顯示導覽
        print("[DEBUG] 準備顯示導覽視窗")
        show_guide(self.root, 'setup.txt', "錯誤碼比對工具導覽")
        print("[DEBUG] 導覽視窗顯示完成")

    def create_widgets(self):
        # 建立大字體 style，並設定 hover 顏色
        style = tb.Style()
        style.configure("Big.TButton", font=("Microsoft JhengHei", 16), padding=10, foreground="#fff", background="#3399ff")
        style.map("Big.TButton",
            background=[("active", "#28a745")],  # 滑鼠移上去的顏色（綠色）
            foreground=[("active", "#fff")]
        )
        style.configure("Main.TFrame", background="#e6f2ff")
        main_frame = tb.Frame(self.root, padding=10, style="Main.TFrame")
        main_frame.pack(fill=BOTH, expand=YES)
        # Error Code XML
        lbl1 = tb.Label(main_frame, text=ui_text['ErrorCodeXMLLabel'])
        lbl1.grid(row=0, column=0, sticky=W, pady=10)
        self.excel1_entry = tb.Entry(main_frame, width=45)
        self.excel1_entry.grid(row=0, column=1, padx=5, pady=10)
        self.browse1_btn = tb.Button(main_frame, text=ui_text['BrowseButton'], bootstyle="outline-primary", command=self.browse_excel1, style="Big.TButton")
        self.browse1_btn.grid(row=0, column=2, padx=5, pady=10)
        ToolTip(self.browse1_btn, ui_text.get('BrowseXMLTooltip', '選擇錯誤碼XML檔案'))
        # Source Excel
        lbl2 = tb.Label(main_frame, text=ui_text['SourceExcelLabel'])
        lbl2.grid(row=1, column=0, sticky=W, pady=10)
        self.excel2_entry = tb.Entry(main_frame, width=45)
        self.excel2_entry.grid(row=1, column=1, padx=5, pady=10)
        self.browse2_btn = tb.Button(main_frame, text=ui_text['BrowseButton'], bootstyle="outline-primary", command=self.browse_excel2, style="Big.TButton")
        self.browse2_btn.grid(row=1, column=2, padx=5, pady=10)
        ToolTip(self.browse2_btn, ui_text.get('BrowseExcelTooltip', '選擇來源Excel檔案'))
        # Sheet select
        lbl3 = tb.Label(main_frame, text=ui_text['SelectSheetLabel'])
        lbl3.grid(row=2, column=0, sticky=W, pady=10)
        self.sheet_combobox = tb.Combobox(main_frame, width=42)
        self.sheet_combobox.grid(row=2, column=1, padx=5, pady=10)
        self.browse3_btn = tb.Button(main_frame, text=ui_text['BrowseButton'], bootstyle="outline-primary", state='disabled', style="Big.TButton")
        self.browse3_btn.grid(row=2, column=2, padx=5, pady=10)
        ToolTip(self.browse3_btn, ui_text.get('BrowseSheetTooltip', '選擇工作表'))
        # Compare button 置中
        self.compare_btn = tb.Button(main_frame, text=ui_text['CompareButton'], bootstyle="outline-success", command=self.compare_files, style="Big.TButton")
        self.compare_btn.grid(row=3, column=0, columnspan=3, pady=25, sticky='ew')
        # 設定所有元件字體
        self.set_all_font_size(self.font_size)

    def set_all_font_size(self, size):
        font = tkfont.Font(size=size, family='Microsoft JhengHei')
        def safe_set_font(widget):
            try:
                if hasattr(widget, 'configure') and 'font' in widget.configure():
                    widget.configure(font=font)
            except Exception:
                pass
        def recursive_set_font(widget):
            safe_set_font(widget)
            for child in widget.winfo_children():
                recursive_set_font(child)
        recursive_set_font(self.root)

    def browse_excel1(self):
        filename = filedialog.askopenfilename(
            title=ui_text['ErrorCodeXMLLabel'],
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.excel1_path = filename
            self.excel1_entry.delete(0, 'end')
            self.excel1_entry.insert(0, filename)

    def browse_excel2(self):
        filename = filedialog.askopenfilename(
            title=ui_text['SourceExcelLabel'],
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.excel2_path = filename
            self.excel2_entry.delete(0, 'end')
            self.excel2_entry.insert(0, filename)
            self.load_sheets()
            # 啟用第三個按鈕
            self.browse3_btn.config(state='normal')

    def load_sheets(self):
        try:
            excel_file = pd.ExcelFile(self.excel2_path)
            sheets = excel_file.sheet_names
            self.sheet_combobox['values'] = sheets
            if sheets:
                self.sheet_combobox.set(sheets[0])
        except Exception as e:
            messagebox.showerror(ui_text['ErrorTitle'], ui_text['SheetLoadError'].format(error=str(e)))

    def compare_files(self):
        if not all([self.excel1_path, self.excel2_path, self.sheet_combobox.get()]):
            messagebox.showerror(ui_text['ErrorTitle'], "Please select all required files and sheet")
            return
        try:
            df_error_codes = pd.read_excel(self.excel1_path, sheet_name="Test Item All")
            df_source = pd.read_excel(self.excel2_path, sheet_name=self.sheet_combobox.get())
            # 自動偵測欄位名稱
            def find_column(df, target):
                for col in df.columns:
                    if str(col).strip().lower() == target.lower():
                        return col
                return None
            desc_col = find_column(df_source, 'Description')
            testid_col = find_column(df_source, 'TestID')
            if not desc_col or not testid_col:
                messagebox.showerror(ui_text['ErrorTitle'], f"找不到 Description 或 TestID 欄位，實際欄位: {df_source.columns.tolist()}")
                return
            df_result = df_source.copy()
            df_result['AB'] = df_source[desc_col]
            df_result['AC'] = df_source[testid_col]
            df_result = df_result[df_result['AB'].notna() & df_result['AC'].notna() & (df_result['AB'].astype(str).str.strip() != '') & (df_result['AC'].astype(str).str.strip() != '')]
            df_result = df_result.reset_index(drop=True)
            output_path = str(Path(self.excel2_path).with_name(f"{Path(self.excel2_path).stem}_compare_ERRORCODE.xlsx"))
            if os.path.exists(output_path):
                overwrite = messagebox.askyesno(ui_text['FileExistsTitle'], ui_text['FileExistsMsg'].format(output_path=output_path))
                if not overwrite:
                    messagebox.showinfo(ui_text['CancelTitle'], ui_text['CancelMsg'])
                    return
            # 只保留 Description 和 TestID 兩欄到 AB
            df_result = df_result[[desc_col, testid_col]].copy()
            df_result.columns = ['你的 description', '你寫的 Error Code']
            # 建立 ErrorCode map（C欄為 TestID，D欄為英文，E欄為中文）
            error_code_map = {str(k).strip(): (str(v1).strip(), str(v2).strip())
                              for k, v1, v2 in zip(df_error_codes.iloc[:, 2], df_error_codes.iloc[:, 3], df_error_codes.iloc[:, 4])}
            cd_list = []
            ce_list = []
            for idx, row in df_result.iterrows():
                test_id = str(row['你寫的 Error Code']).strip()
                if test_id in error_code_map:
                    description, chinese_desc = error_code_map[test_id]
                    cd_list.append(description)
                    ce_list.append(chinese_desc)
                else:
                    cd_list.append(ui_text['NotFound'])
                    ce_list.append(ui_text['NotFoundCN'])
            df_result['Test Item 文件的 description'] = cd_list
            df_result['Test Item 的 Error Code'] = ce_list
            # 明確指定欄位順序
            df_result = df_result[['你的 description', '你寫的 Error Code', 'Test Item 文件的 description', 'Test Item 的 Error Code']]
            # 將結果寫入第一個 sheet，來源 Test Item All 完整複製到 sheet2
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name=self.sheet_combobox.get())
                df_error_codes.to_excel(writer, index=False, sheet_name='Test Item All')
            # 用 openpyxl 美化 Excel：字型 Calibri、標題加粗、加框線、自動欄寬
            wb = load_workbook(output_path)
            # 美化所有 sheet
            for ws in wb.worksheets:
                calibri_font = Font(name='Calibri', size=11)
                bold_font = Font(name='Calibri', size=11, bold=True)
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                # 標題列加粗
                for cell in ws[1]:
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                # 內容列設字型、框線
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = calibri_font
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                # 自動調整欄寬
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
            # 反白 Test Item All sheet 比對到的 TestID
            try:
                ws2 = wb['Test Item All']
                # 來源 TestID set
                source_testids = set(str(tid).strip() for tid in df_result['AC'])
                highlight_fill = PatternFill("solid", fgColor="00BFFF")  # 藍色
                for row in ws2.iter_rows(min_row=2):  # 跳過標題
                    testid_cell = row[2]  # C欄
                    if str(testid_cell.value).strip() in source_testids:
                        for cell in row:
                            cell.fill = highlight_fill
            except Exception as e:
                pass  # 若 sheet 名稱或格式異常不影響主流程
            wb.save(output_path)
            messagebox.showinfo(ui_text['SuccessTitle'], f"{ui_text['SuccessMsg']}\n{output_path}")
        except Exception as e:
            messagebox.showerror(ui_text['ErrorTitle'], ui_text['CompareError'].format(error=str(e)))

    def center_window(self):
        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        ws = self.root.winfo_screenwidth()
        hs = self.root.winfo_screenheight()
        x = (ws // 2) - (w // 2)
        y = (hs // 2) - (h // 2)
        self.root.geometry(f'+{x}+{y}')

if __name__ == "__main__":
    import tkinter as tk
    from guide_popup.guide import show_guide

    root = tk.Tk()
    app = ErrorCodeComparer(root)
    root.mainloop() 