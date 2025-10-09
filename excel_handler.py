"""
Excel處理模組
負責處理所有Excel相關的操作，包括讀取、比對和寫入Excel檔案
"""
import os
import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from typing import Tuple, Dict, Optional

logger = logging.getLogger(__name__)

class ExcelHandler:
    """Excel 檔案處理類別，負責讀取、比對、寫入、格式化等操作"""
    def __init__(self):
        self.error_code_map: Dict[str, Tuple[str, str]] = {}
        self.current_sheet: Optional[str] = None

    def load_error_codes(self, file_path: str) -> bool:
        """載入錯誤碼Excel檔案，建立 TestID 對應說明的字典"""
        try:
            df_error_codes = pd.read_excel(file_path, sheet_name="Test Item All")
            self.error_code_map = {
                str(k).strip(): (str(v1).strip(), str(v2).strip())
                for k, v1, v2 in zip(df_error_codes.iloc[:, 2], df_error_codes.iloc[:, 3], df_error_codes.iloc[:, 4])
            }
            logger.info(f"成功載入錯誤碼檔案: {file_path}")
            return True
        except Exception as e:
            logger.error(f"載入錯誤碼檔案時發生錯誤: {str(e)}")
            return False

    def load_source_sheet(self, file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """載入來源Excel檔案的指定工作表"""
        try:
            # 嘗試不同的讀取方式來處理標題行
            try:
                # 首先嘗試正常讀取
                df_source = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # 檢查是否有 "Unnamed" 欄位，如果有則嘗試跳過第一行
                if any('Unnamed' in str(col) for col in df_source.columns):
                    logger.info("檢測到 Unnamed 欄位，嘗試跳過第一行重新讀取")
                    df_source = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=1)
                    
                    # 如果還是有 Unnamed，嘗試跳過更多行
                    if any('Unnamed' in str(col) for col in df_source.columns):
                        logger.info("仍有 Unnamed 欄位，嘗試跳過前兩行")
                        df_source = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)
                
                # 如果還是有 Unnamed，嘗試使用 header=None 並手動設定欄位名稱
                if any('Unnamed' in str(col) for col in df_source.columns):
                    logger.info("仍有 Unnamed 欄位，嘗試使用 header=None")
                    df_source = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    # 尋找包含 "Main Function" 的行作為標題行
                    header_row = None
                    for i, row in df_source.iterrows():
                        if 'Main Function' in str(row.values):
                            header_row = i
                            break
                    
                    if header_row is not None:
                        logger.info(f"找到標題行在第 {header_row + 1} 行")
                        df_source.columns = df_source.iloc[header_row]
                        df_source = df_source.drop(df_source.index[:header_row + 1]).reset_index(drop=True)
                    else:
                        # 如果找不到標題行，使用第一行
                        df_source.columns = df_source.iloc[0]
                        df_source = df_source.drop(df_source.index[0]).reset_index(drop=True)
                
                # 清理欄位名稱
                df_source.columns = [str(col).strip() for col in df_source.columns]
                
            except Exception as e:
                logger.warning(f"特殊讀取方式失敗，使用預設方式: {e}")
                df_source = pd.read_excel(file_path, sheet_name=sheet_name)
            
            self.current_sheet = sheet_name
            logger.info(f"成功載入來源工作表: {sheet_name}")
            logger.info(f"欄位名稱: {list(df_source.columns)}")
            return df_source
        except Exception as e:
            logger.error(f"載入來源工作表時發生錯誤: {str(e)}")
            return None

    def find_column(self, df: pd.DataFrame, target: str) -> Optional[str]:
        """在DataFrame中尋找目標欄位名稱（忽略大小寫與空白）"""
        for col in df.columns:
            if str(col).strip().lower() == target.lower():
                return col
        return None

    def compare_data(self, df_source: pd.DataFrame, not_found_text: str, not_found_cn_text: str) -> Optional[pd.DataFrame]:
        """比對來源資料與錯誤碼字典，產生比對結果DataFrame（已優化可用merge）"""
        try:
            desc_col = self.find_column(df_source, 'Description')
            testid_col = self.find_column(df_source, 'TestID')
            if not desc_col or not testid_col:
                logger.error(f"找不到必要欄位，實際欄位: {df_source.columns.tolist()}")
                return None
            df_result = df_source[[desc_col, testid_col]].copy()
            # 修改 header 名稱
            df_result.columns = ['你的 description', '你寫的 Error Code']
            # merge 或 for 迴圈比對
            cd_list = []
            ce_list = []
            for idx, row in df_result.iterrows():
                test_id = str(row['你寫的 Error Code']).strip()
                if test_id in self.error_code_map:
                    description, chinese_desc = self.error_code_map[test_id]
                    cd_list.append(description)
                    ce_list.append(chinese_desc)
                else:
                    cd_list.append(not_found_text)
                    ce_list.append(not_found_cn_text)
            df_result['Test Item 文件的 description'] = cd_list
            df_result['Test Item 的 Error Code'] = ce_list
            logger.info("成功完成資料比對")
            return df_result
        except Exception as e:
            logger.error(f"比對資料時發生錯誤: {str(e)}")
            return None

    def save_result(self, df_result: pd.DataFrame, df_error_codes: pd.DataFrame, 
                   output_path: str, sheet_name: str, ai_recommendations: list = None) -> bool:
        """儲存比對結果，並反白來源TestID對應Test Item All行"""
        try:
            # 檢查輸出檔案是否被佔用
            if os.path.exists(output_path):
                try:
                    # 嘗試開啟檔案檢查是否被佔用
                    with open(output_path, 'a'):
                        pass
                except PermissionError:
                    logger.warning(f"檔案被佔用，嘗試重新命名: {output_path}")
                    # 生成備用檔名
                    base_name = os.path.splitext(output_path)[0]
                    extension = os.path.splitext(output_path)[1]
                    counter = 1
                    while True:
                        new_path = f"{base_name}_backup_{counter}{extension}"
                        try:
                            with open(new_path, 'a'):
                                pass
                            output_path = new_path
                            logger.info(f"使用備用檔名: {output_path}")
                            break
                        except PermissionError:
                            counter += 1
                            if counter > 10:  # 避免無限循環
                                raise Exception(f"無法找到可用的檔名，檔案可能被多個程式佔用")
            
            # 如果有 AI 推薦，新增 E、F 欄位
            if ai_recommendations and len(ai_recommendations) > 0:
                df_result = self._add_ai_recommendations(df_result, ai_recommendations)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name=sheet_name)
                df_error_codes.to_excel(writer, index=False, sheet_name='Test Item All')
            # highlight_testids: 來源TestID
            highlight_testids = [str(tid).strip() for tid in df_result['你寫的 Error Code']]
            self._format_excel(output_path, highlight_testids=highlight_testids)
            logger.info(f"成功儲存比對結果: {output_path}")
            return True
        except Exception as e:
            logger.error(f"儲存比對結果時發生錯誤: {str(e)}")
            return False

    def _add_ai_recommendations(self, df_result: pd.DataFrame, ai_recommendations: list) -> pd.DataFrame:
        """
        為 DataFrame 新增 AI 推薦的 E、F 欄位
        
        Args:
            df_result: 原始結果 DataFrame
            ai_recommendations: AI 推薦列表，格式為 [(test_id_1, test_id_2), ...]
            
        Returns:
            pd.DataFrame: 新增 E、F 欄位後的 DataFrame
        """
        try:
            # 確保 ai_recommendations 長度與 DataFrame 行數一致
            if len(ai_recommendations) != len(df_result):
                logger.warning(f"AI 推薦數量 ({len(ai_recommendations)}) 與 DataFrame 行數 ({len(df_result)}) 不一致")
                # 補齊不足的推薦
                while len(ai_recommendations) < len(df_result):
                    ai_recommendations.append(("", ""))
                # 截斷多餘的推薦
                ai_recommendations = ai_recommendations[:len(df_result)]
            
            # 新增 E、F 欄位
            df_result = df_result.copy()
            df_result['AI推薦 test ID 1'] = [rec[0] for rec in ai_recommendations]
            df_result['AI推薦 test ID 2'] = [rec[1] for rec in ai_recommendations]
            
            logger.info(f"成功新增 AI 推薦欄位，共 {len(ai_recommendations)} 個推薦")
            return df_result
        except Exception as e:
            logger.error(f"新增 AI 推薦欄位時發生錯誤: {str(e)}")
            return df_result

    def add_ai_recommendations_to_existing_file(self, file_path: str, ai_recommendations: list) -> bool:
        """
        為現有的比對結果檔案新增 AI 推薦欄位
        使用 openpyxl 直接操作，完全保持原有格式（包括字體和行高）
        
        Args:
            file_path: 現有檔案路徑
            ai_recommendations: AI 推薦列表，格式為 [(test_id, chinese_desc), ...]
            
        Returns:
            bool: 是否成功
        """
        try:
            # 使用 openpyxl 直接載入工作簿，保持所有格式
            wb = load_workbook(file_path)
            ws = wb.active  # 使用第一個工作表
            
            logger.info(f"使用 openpyxl 載入檔案: {file_path}")
            logger.info(f"工作表名稱: {ws.title}")
            logger.info(f"工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
            
            # 確保 ai_recommendations 長度與資料行數一致
            data_rows = ws.max_row - 1  # 減去標題行
            if len(ai_recommendations) != data_rows:
                logger.warning(f"AI 推薦數量 ({len(ai_recommendations)}) 與資料行數 ({data_rows}) 不一致")
                # 補齊不足的推薦
                while len(ai_recommendations) < data_rows:
                    ai_recommendations.append(("", ""))
                # 截斷多餘的推薦
                ai_recommendations = ai_recommendations[:data_rows]
            
            # 檢查是否已經有 AI 推薦欄位
            col_e_exists = False
            col_g_exists = False
            col_e_index = None
            col_g_index = None
            
            # 檢查 E 和 G 列是否已經有 AI 推薦欄位
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value and 'AI推薦 test ID' in str(cell_value):
                    col_e_exists = True
                    col_e_index = col_idx
                elif cell_value and 'AI推薦 中文' in str(cell_value):
                    col_g_exists = True
                    col_g_index = col_idx
            
            # 如果沒有 AI 推薦欄位，新增到 E 和 G 列
            if not col_e_exists:
                col_e_index = ws.max_column + 1
                ws.cell(row=1, column=col_e_index, value='AI推薦 test ID')
                logger.info(f"新增 AI推薦 test ID 欄位到第 {col_e_index} 列")
            
            if not col_g_exists:
                col_g_index = ws.max_column + 1
                ws.cell(row=1, column=col_g_index, value='AI推薦 中文')
                logger.info(f"新增 AI推薦 中文 欄位到第 {col_g_index} 列")
            
            # 寫入 AI 推薦資料，保持原有格式
            for row_idx, (test_id, chinese_desc) in enumerate(ai_recommendations, start=2):  # 從第2行開始（跳過標題）
                if col_e_index:
                    cell_e = ws.cell(row=row_idx, column=col_e_index, value=test_id)
                    # 保持與其他資料行相同的格式
                    self._apply_data_cell_format(cell_e)
                
                if col_g_index:
                    cell_g = ws.cell(row=row_idx, column=col_g_index, value=chinese_desc)
                    # 保持與其他資料行相同的格式
                    self._apply_data_cell_format(cell_g)
            
            # 如果新增了欄位，需要為標題行應用格式
            if not col_e_exists or not col_g_exists:
                self._apply_header_format(ws, col_e_index if not col_e_exists else None)
                self._apply_header_format(ws, col_g_index if not col_g_exists else None)
            
            # 自動調整 AI 推薦欄位的寬度，確保內容完整顯示
            self._auto_adjust_column_widths(ws, col_e_index, col_g_index)
            
            # 保存檔案，保持所有原有格式
            wb.save(file_path)
            wb.close()
            
            logger.info(f"成功為現有檔案新增 AI 推薦欄位: {file_path}")
            logger.info(f"AI推薦 test ID 寫入到第 {col_e_index} 列")
            logger.info(f"AI推薦 中文 寫入到第 {col_g_index} 列")
            logger.info(f"共寫入 {len(ai_recommendations)} 筆推薦資料")
            
            return True
        except Exception as e:
            logger.error(f"為現有檔案新增 AI 推薦欄位時發生錯誤: {str(e)}")
            return False
    
    def _apply_data_cell_format(self, cell):
        """為資料儲存格應用標準格式（與比對結果一致）"""
        from openpyxl.styles import Font, Border, Side
        
        # 設定字體為 Calibri 12
        cell.font = Font(name="Calibri", size=12)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        cell.border = thin_border
    
    def _auto_adjust_column_widths(self, worksheet, col_e_index, col_g_index):
        """
        自動調整 AI 推薦欄位的寬度，確保內容完整顯示
        
        Args:
            worksheet: 工作表物件
            col_e_index: E 欄位索引
            col_g_index: G 欄位索引
        """
        try:
            # 定義各欄位的建議寬度
            column_widths = {
                col_e_index: 15,  # AI推薦 test ID - Test ID 通常較短
                col_g_index: 25   # AI推薦 中文 - 中文描述需要更多空間
            }
            
            # 為每個欄位設定寬度
            for col_index, width in column_widths.items():
                if col_index:
                    worksheet.column_dimensions[worksheet.cell(row=1, column=col_index).column_letter].width = width
                    logger.info(f"設定第 {col_index} 欄位寬度為 {width}")
            
            # 額外檢查：如果內容超過設定寬度，動態調整
            for col_index in [col_e_index, col_g_index]:
                if col_index:
                    max_length = 0
                    col_letter = worksheet.cell(row=1, column=col_index).column_letter
                    
                    # 檢查標題和所有資料行的長度
                    for row_idx in range(1, worksheet.max_row + 1):
                        cell_value = worksheet.cell(row=row_idx, column=col_index).value
                        if cell_value:
                            # 計算字串長度（中文字算2個字元）
                            str_length = len(str(cell_value))
                            # 中文字符長度調整
                            chinese_chars = sum(1 for char in str(cell_value) if '\u4e00' <= char <= '\u9fff')
                            adjusted_length = str_length + chinese_chars
                            max_length = max(max_length, adjusted_length)
                    
                    # 設定最小寬度為 12，最大寬度為 50
                    optimal_width = max(12, min(50, max_length + 2))
                    worksheet.column_dimensions[col_letter].width = optimal_width
                    logger.info(f"動態調整第 {col_index} 欄位寬度為 {optimal_width} (內容最大長度: {max_length})")
            
        except Exception as e:
            logger.error(f"自動調整欄位寬度時發生錯誤: {str(e)}")
    
    def _apply_header_format(self, worksheet, col_index):
        """為標題儲存格應用標準格式（與比對結果一致）"""
        if col_index is None:
            return
            
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        cell = worksheet.cell(row=1, column=col_index)
        
        # 設定標題格式（綠色背景，粗體，置中）
        green_fill = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
        bold_font = Font(name="Calibri", size=12, bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        cell.fill = green_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    def _format_excel(self, file_path: str, highlight_testids: list = None):
        """格式化Excel檔案，並反白 Test Item All sheet 對應 TestID 行"""
        try:
            wb = load_workbook(file_path)
            calibri_font = Font(name='Calibri', size=12)
            bold_font = Font(name='Calibri', size=12, bold=True)
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            green_fill = PatternFill("solid", fgColor="00C853")  # 綠色
            for ws in wb.worksheets:
                # 設定標題列樣式
                for cell in ws[1]:
                    cell.font = bold_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = green_fill
                # 設定資料列樣式
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = calibri_font
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                # 自動調整欄寬（優化：考慮中文字、英文、數字）
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                cell_str = str(cell.value)
                                cell_len = len(cell_str)
                                # 中文字符長度調整（中文字算2個字元）
                                chinese_chars = sum(1 for char in cell_str if '\u4e00' <= char <= '\u9fff')
                                adjusted_len = cell_len + chinese_chars
                                max_length = max(max_length, adjusted_len)
                        except:
                            pass
                    # 設定最小寬度為 12，最大寬度為 50
                    optimal_width = max(12, min(50, max_length + 2))
                    ws.column_dimensions[column_letter].width = optimal_width
                # 凍結第一列
                ws.freeze_panes = ws["A2"]
            # 反白 Test Item All sheet 對應 TestID 行
            if highlight_testids:
                try:
                    ws2 = wb['Test Item All']
                    for row in ws2.iter_rows(min_row=2):  # 跳過標題
                        testid_cell = row[2]  # C欄
                        if str(testid_cell.value).strip() in highlight_testids:
                            for cell in row:
                                cell.fill = green_fill
                except Exception as e:
                    pass  # 若 sheet 名稱或格式異常不影響主流程
            wb.save(file_path)
            logger.info("成功格式化Excel檔案")
        except Exception as e:
            logger.error(f"格式化Excel檔案時發生錯誤: {str(e)}")

    def get_sheet_names(self, file_path: str) -> list:
        """獲取Excel檔案中的所有工作表名稱"""
        try:
            excel_file = pd.ExcelFile(file_path)
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"獲取工作表名稱時發生錯誤: {str(e)}")
            return [] 